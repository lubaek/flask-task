from flask import Flask, render_template, request
from openpyxl import load_workbook


app = Flask(__name__)


@app.route('/')
def homepage():
    excel = load_workbook('my_report.xlsx')
    sheet = excel['Sheet1']
    column = sheet['A']
    return render_template('index.html', goods=column)


@app.route('/add/', methods=["POST"])
def add():
    good = request.form["good"]
    excel = load_workbook('my_report.xlsx')
    sheet = excel['Sheet1']
    sheet.cell(row=sheet.max_row+1, column=1).value = good
    print(sheet.max_row+1)
    excel.save('my_report.xlsx')
    return """
        <h3>Инвертарь пополнен</h3>
        <a href="/">Домой</a>
    """


# if __name__ == "__main__":
#     app.run(debug=True)
